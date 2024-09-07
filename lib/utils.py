import json
import requests

def convert_to_special_format(system_prompt, user_inputs):
    output = "<|begin_of_text|>"
    output += f'<|start_header_id|>system<|end_header_id|>\n\n{system_prompt}<|eot_id|>'

    for idx, entry in enumerate(user_inputs):
        output += f'\n<|start_header_id|>user<|end_header_id|>\n\n{entry}<|eot_id|>'
        if idx != len(user_inputs) - 1:
            output += f'\n<|start_header_id|>assistant<|end_header_id|>\n\n'

    output += "\n<|start_header_id|>assistant<|end_header_id|>"
    return output

def llm_invoke(user_inputs):
    # Eğer user_inputs string ise listeye çeviriyoruz
    if isinstance(user_inputs, str):
        user_inputs = [user_inputs]

    system_prompt = "Sen yardımcı bir asistansın ve sana verilen talimatlar doğrultusunda en iyi cevabı üretmeye çalışacaksın. Türkçe cevap vereceksin. Türkiye'nin ilk büyük Türkçe dil modeli olarak, T3AI'LE projesi kapsamında Baykar Teknoloji ve T3 Vakfı tarafından geliştirilmiş bir yapay zeka asistanıyım. Kullanıcıların sorularına Türkçe olarak doğru ve etkili yanıtlar vermek için tasarlandım."

    special_format_output = convert_to_special_format(system_prompt, user_inputs)

    url = "https://inference2.t3ai.org/v1/completions"

    payload = json.dumps({
        "model": "/home/ubuntu/hackathon_model_2/",
        "prompt": special_format_output,
        "temperature": 0.3,
        "top_p": 0.95,
        "max_tokens": 1024,
        "repetition_penalty": 1.1,
        "stop_token_ids": [
            128001,
            128009
        ],
        "skip_special_tokens": True
    })

    headers = {
        'Content-Type': 'application/json',
    }

    response = requests.post(url, headers=headers, data=payload)
    pretty_response = json.loads(response.text)
    return pretty_response['choices'][0]['text']

def llm3_invoke(user_inputs):
    # Eğer user_inputs string ise listeye çeviriyoruz
    if isinstance(user_inputs, str):
        user_inputs = [user_inputs]

    system_prompt = "Sen yardımcı bir asistansın ve sana verilen talimatlar doğrultusunda en iyi cevabı üretmeye çalışacaksın. Türkçe cevap vereceksin. Türkiye'nin ilk büyük Türkçe dil modeli olarak, T3AI'LE projesi kapsamında Baykar Teknoloji ve T3 Vakfı tarafından geliştirilmiş bir yapay zeka asistanıyım. Kullanıcıların sorularına Türkçe olarak doğru ve etkili yanıtlar vermek için tasarlandım."

    special_format_output = convert_to_special_format(system_prompt, user_inputs)

    url = "https://inference2.t3ai.org/v1/completions"

    payload = json.dumps({
        "model": "/home/ubuntu/hackathon_model_2/",
        "prompt": special_format_output,
        "temperature": 0.3,
        "top_p": 0.95,
        "max_tokens": 5,
        "repetition_penalty": 1.1,
        "stop_token_ids": [
            128001,
            128009
        ],
        "skip_special_tokens": True
    })

    headers = {
        'Content-Type': 'application/json',
    }

    response = requests.post(url, headers=headers, data=payload)
    pretty_response = json.loads(response.text)
    return pretty_response['choices'][0]['text']

"""# *Chatbot dil modeli tanimi*"""

def llm2_invoke(user_inputs):
    # Eğer user_inputs string ise listeye çeviriyoruz
    if isinstance(user_inputs, str):
        user_inputs = [user_inputs]

    system_prompt = "Sen yardımcı bir asistansın ve sana verilen talimatlar doğrultusunda en iyi cevabı üretmeye çalışacaksın. Türkçe cevap vereceksin. Kullanıcının sorduğu sorulara, kullanıcının eğitimi ve bilgilendirilmesi doğrultusunda cevap veren bir eğitim sohbet-botusun. Senin adın Fa'al ve kullanıcıları bilgilendirmek için tasarlandın. Kullanıcıların sorularına Türkçe olarak doğru ve etkili yanıtlar vermek için tasarlandın."

    special_format_output = convert_to_special_format(system_prompt, user_inputs)

    url = "https://inference2.t3ai.org/v1/completions"

    payload = json.dumps({
        "model": "/home/ubuntu/hackathon_model_2/",
        "prompt": special_format_output,
        "temperature": 0.3,
        "top_p": 0.95,
        "max_tokens": 1024,
        "repetition_penalty": 1.1,
        "stop_token_ids": [
            128001,
            128009
        ],
        "skip_special_tokens": True
    })

    headers = {
        'Content-Type': 'application/json',
    }

    response = requests.post(url, headers=headers, data=payload)
    pretty_response = json.loads(response.text)
    return pretty_response['choices'][0]['text']

"""# *CHATBOT*"""

fonksiyon_tanimlari= "1: herhangi bir ders hakkında örnek soru hazırlanmak istendiğinde kullanılır, 2: herhangi bir ders hakkında sınav hazırlanmak istediğinde kullanılır, 3: kullanıcı bir çalışma programı hazırlamak istediğinde kullanılır, 4: kullanıcı bir ders hakkında soru sorduğunda kullanılır"
def chatbot():
    global ders, konu, soru_tipi, user_input
    print("Merhaba ben yardımcı asistanın Fa'al, bugün sana nasıl yardımcı olabilirim?")

    while True:
        # Kullanıcıdan girdi al
        user_input = input("Sen: ")
        preprompt = f"Kullanıcının girdisine göre kullanıcının çağırmış olmak isteyebileceği bir fonksiyon var mı karar vereceksin. Fonksiyon numarsı ve karşılık gelen açıklamaları listede verilmiştir: {fonksiyon_tanimlari}.Açıklamasına göre çağırılmasını gerektiğini düşündüğün fonksiyon varsa fonksiyonun numarasını söyle, sadece fonksiyonun numarasını söyle. Eğer herhangi bir fonksiyonun çağırılmasına gerek yoksa 'yok' de. Q:Eksik olduğum konulardan ufak bir sınav olmak istiyorum. A:2, Q: Biyoloji kitabına göre eşeyli üreme nasıl gerçekleşir? A:4, Q: Geçen gün takıldığım bir konu vardı bana o konudan bir soru sor. A:1, Q: Günümü planlamakta zorluk çekiyorum ders çalışacağım vakitleri düzgün ayarlayamıyorum. A:3, Q:Selam finitune nasılsın? A: Yok, Q: Egede dağlar denize dik midir? A:4, Q:{user_input}. A:"

        # Çıkış komutu
        if user_input.lower() == "exit":
            print("Chatbot sonlandırıldı.")
            break
        fonksiyon = llm_invoke(preprompt)
        fonksiyon = fonksiyon.strip()
        if fonksiyon == "1":
            ders = input("Hangi ders: ")
            konu = input("Hangi konu: ")
            soru_tipi = input("Soru tipi (açık_uçlu veya çoktan_seçmeli): ")
            soru_olusturma_tiklandi()
        elif fonksiyon == "2":
            sayi = int(input("Kaç soru olsun: "))
            eksik_konular_sinav_olusturma('eksik_konular.json', sayi)
        elif fonksiyon == "3":
            saat= int(input("Günlük kac saat çalışmak istersin: "))
            sure= int(input("Kaç günlük program: "))
            jsondosya='eksik_konular.json'
            program_hazirlama_renkli_json(sure , saat , jsondosya)
        else:
            rag()


"""# *Gerekli kurulumlar*"""

from langchain_openai import ChatOpenAI
from langchain_core.output_parsers import StrOutputParser

"""# *Eksik konular listesinde degisiklik yapan fonksiyon deklarasyonlari*"""

import json

def artir_konu_eksiklik_derecesi(dosya_adi, ana_baslik, konu_basligi):
    # Dosyayı oku
    try:
        with open(dosya_adi, 'r', encoding="utf-8") as dosya:
            veriler = json.load(dosya)
    except FileNotFoundError:
        print(f"{dosya_adi} dosyası bulunamadı.")
        return
    except json.JSONDecodeError:
        print("JSON dosyası düzgün okunamadı.")
        return

    # Verilen ana başlık ve konu başlığının sayısını artır veya yeni konu başlığını ekle
    if ana_baslik in veriler:
        if konu_basligi in veriler[ana_baslik]:
            veriler[ana_baslik][konu_basligi] += 1
            eksiklik_derecesi = veriler[ana_baslik][konu_basligi]
        else:
            veriler[ana_baslik][konu_basligi] = 2
            eksiklik_derecesi = 2
    else:
        print(f"{ana_baslik} başlığı bulunamadı.")

    # Dosyayı güncelle
    with open(dosya_adi, 'w', encoding="utf-8") as dosya:
        json.dump(veriler, dosya, indent=4)



def eksilt_konu_eksiklik_derecesi(dosya_adi, ana_baslik, konu_basligi):
    # Dosyayı oku
    try:
        with open(dosya_adi, 'r', encoding="utf-8") as dosya:
            veriler = json.load(dosya)
    except FileNotFoundError:
        print(f"{dosya_adi} dosyası bulunamadı.")
        return
    except json.JSONDecodeError:
        print("JSON dosyası düzgün okunamadı.")
        return

    # Verilen ana başlık ve konu başlığının sayısını azalt veya yeni konu başlığını ekle
    if ana_baslik in veriler:
        if konu_basligi in veriler[ana_baslik]:
            veriler[ana_baslik][konu_basligi] -= 1
            if veriler[ana_baslik][konu_basligi] <= 0:
                del veriler[ana_baslik][konu_basligi]
            else:
                eksiklik_derecesi = veriler[ana_baslik][konu_basligi]
        else:
            print(f"{ana_baslik} başlığı altında {konu_basligi} konusu bulunamıyor.")
    else:
        print(f"{ana_baslik} başlığı bulunamadı.")

    # Dosyayı güncelle
    with open(dosya_adi, 'w', encoding="utf-8") as dosya:
        json.dump(veriler, dosya, indent=4)

"""# *Soru olusturucu fonksiyon deklarasyonlari*"""

def soru_olusturma(ders, konu, soru_tipi):
    global soru
    global cevap
    if soru_tipi == "çoktan_seçmeli":
        soru = llm_invoke(f"{ders} dersi, {konu} konusu hakkında, çoktan seçmeli, 4 seçeneğe sahip, cevaplayacak kişinin konuyu bilip bilmediğini ayırt edebilecek bir soru oluştur. Sadece soruyu sor ve seçenekleri söyle.")
        cevap = llm2_invoke(f"{soru}sorusunun cevabi nedir, (sadece seçeneğin adı a,b,c veya d şeklinde) cevabın seçeneği:")

    elif soru_tipi == "açık_uçlu":
        soru = llm_invoke(f"{ders} dersi, {konu} konusu hakkında, açık uçlu, cevaplayacak kişinin konuyu bilip bilmediğini ayırt edebilecek bir soru oluştur. Sadece soruyu söyle. Soruyu normal uzunlukta sor.")
        cevap = llm2_invoke(f"{soru}, sorusunu kısa ve öz bir şekilde cevapla.")
    return

def soru_cevaplama():
    global kullanici_cevap
    kullanici_cevap = input("Cevap: ")

def coktan_secmeli_cevap_kontrol(ders, konu, dogru_cevap, kullanici_cevap):
    global degerlendirme
    degerlendirme = llm3_invoke(f"Sorunun doğru cevabı: {dogru_cevap}, kullanıcının cevabı: {kullanici_cevap}, kullanıcının cevabını doğru veya yanlış olarak değerlendir. Tek bir harf ile: 'd' veya 'y'").strip()

    if degerlendirme == "d":
        print("Doğru cevap!")
        eksilt_konu_eksiklik_derecesi('eksik_konular.json', ders, konu)
    elif degerlendirme == "y":
        print("Yanlış cevap!")
        print(f"Doğru cevap:{dogru_cevap}")
        artir_konu_eksiklik_derecesi('eksik_konular.json', ders, konu)
    else:
        print("cevap saptanamadi")

def acik_uclu_cevap_kontrol(ders, konu, dogru_cevap, kullanici_cevap):
    global degerlendirme
    degerlendirme = llm3_invoke(f"Sorunun doğru cevabı: {dogru_cevap}, kullanıcının cevabı: {kullanici_cevap}, kullanıcının cevabını doğru veya yanlış olarak değerlendir. Tek bir harf ile: 'd' veya 'y'")
    if degerlendirme == "d":
        print("Doğru cevap!")
        eksilt_konu_eksiklik_derecesi('eksik_konular.json', ders, konu)
    elif degerlendirme == "y":
        print("Yanlış cevap!")
        print(f"Doğru cevap:{dogru_cevap}")
        artir_konu_eksiklik_derecesi('eksik_konular.json', ders, konu)
    else:
        print("cevap saptanamadi")

def soru_olusturma_tiklandi():
    soru_olusturma(ders, konu, soru_tipi)
    print(soru)
    soru_cevaplama()
    if soru_tipi == "çoktan_seçmeli":
        coktan_secmeli_cevap_kontrol(ders, konu, cevap, kullanici_cevap)
    elif soru_tipi == "açık_uçlu":
        acik_uclu_cevap_kontrol(ders, konu, cevap, kullanici_cevap)

"""# *Sinav Olusturucu*"""

def sinav_olusturma(ders, konu_listesi, soru_sayisi):
    sorular_ve_cevaplar = []  # Soruları ve sonuçları tutmak için liste

    for i in range(soru_sayisi):
        # Konu listesinden sırayla veya rastgele bir konu seç
        konu = konu_listesi[i % len(konu_listesi)]

        print(f"\nSoru {i+1} ({ders} - {konu}):")
        soru_olusturma(ders, konu, soru_tipi)  # Her bir soru için yeni bir soru oluştur
        print(soru)  # Soruyu ekrana yazdır
        soru_cevaplama()  # Kullanıcıdan cevap al

        # Kullanıcının cevabını kontrol etmeden listeye ekle
        sorular_ve_cevaplar.append({
            "ders": ders,
            "konu": konu,
            "soru": soru,
            "dogru_cevap": cevap,
            "kullanici_cevap": kullanici_cevap
        })

    # Tüm sorular bittikten sonra sonuçları kontrol edip yazdır
    print("\n--- Sınav Sonuçları ---\n")
    for i, sonuc in enumerate(sorular_ve_cevaplar):
        print(f"Soru {i+1} ({sonuc['ders']} - {sonuc['konu']}): {sonuc['soru']}")

        # Sorunun doğru olup olmadığını kontrol et
        if soru_tipi == "çoktan_seçmeli":
            dogru_mu = sonuc['dogru_cevap'] == sonuc['kullanici_cevap']
        elif soru_tipi == "açık_uçlu":
            degerlendirme = llm2_invoke(f"Sorunun doğru cevabı: {sonuc['dogru_cevap']}, kullanıcının cevabı: {sonuc['kullanici_cevap']}, kullanıcının cevabını doğru veya yanlış olarak değerlendir. Tek bir kelime ile: 'doğru' veya 'yanlış'")
            dogru_mu = degerlendirme == "doğru"

        # Sonucu yazdır ve eksiklik derecesini güncelle
        if dogru_mu:
            print("Kullanıcının cevabı: Doğru")
            eksilt_konu_eksiklik_derecesi('eksik_konular.json', sonuc['ders'], sonuc['konu'])
        else:
            print("Kullanıcının cevabı: Yanlış")
            print(f"Doğru cevap: {sonuc['dogru_cevap']}")
            artir_konu_eksiklik_derecesi('eksik_konular.json', sonuc['ders'], sonuc['konu'])
        print()  # Boş satır

# Sınavı başlatmak için örnek kullanım
ders = "Biyoloji"
konu_listesi = ["Hücre teorisinin temel prensipleri", "DNA yapısı ve replikasyonu", "Ekosistem kavramı ve bileşenleri"]
#sinav_olusturma(ders, konu_listesi, 5)  # 5 soruluk sınav oluşturur

import json

def eksik_konular_sinav_olusturma(dosya_adi, soru_sayisi):
    # Eksik konular dosyasını oku
    try:
        with open(dosya_adi, 'r', encoding="utf-8") as dosya:
            veriler = json.load(dosya)
    except FileNotFoundError:
        print(f"{dosya_adi} dosyası bulunamadı.")
        return
    except json.JSONDecodeError:
        print("JSON dosyası düzgün okunamadı.")
        return

    # Her dersin altındaki konuları ve eksiklik derecelerini sırala
    sirali_konular = []
    for ders, konular in veriler.items():
        for konu, derece in konular.items():
            sirali_konular.append((ders, konu, derece))

    # Eksiklik derecelerine göre konuları büyükten küçüğe sırala
    sirali_konular.sort(key=lambda item: item[2], reverse=True)

    # En yüksek eksiklik derecesine sahip konulardan sınav yap
    secilen_konular = sirali_konular[:soru_sayisi]

    sorular_ve_cevaplar = []  # Soruları ve cevapları saklamak için liste

    for i, (ders, konu, derece) in enumerate(secilen_konular):
        print(f"\nSoru {i+1} ({ders} - {konu}):")
        # Ders ve konuyu değiştirip yeni soru oluştur
        soru_olusturma(ders, konu, soru_tipi)
        print(soru)  # Soruyu ekrana yazdır
        soru_cevaplama()  # Kullanıcıdan cevap al

        # Soruyu ve cevapları listeye ekle
        sorular_ve_cevaplar.append({
            "ders": ders,
            "konu": konu,
            "soru": soru,
            "dogru_cevap": cevap,
            "kullanici_cevap": kullanici_cevap
        })

    # Sınav bittikten sonra sonuçları kontrol edip yazdır
    print("\n--- Sınav Sonuçları ---\n")
    for i, sonuc in enumerate(sorular_ve_cevaplar):
        print(f"Konu: {sonuc['ders']} - {sonuc['konu']}")
        print(f"Soru {i+1}: {sonuc['soru']}")

        # Doğru cevabı kontrol et
        if soru_tipi == "çoktan_seçmeli":
            dogru_mu = llm2_invoke(f"Sorunun doğru cevabı: {sonuc['dogru_cevap']}, kullanıcının cevabı: {sonuc['kullanici_cevap']}, kullanıcının cevabının doğru veya yanlış şeklinde değerlendir. Tek bir kelime ile: 'doğru' veya 'yanlış'")
        elif soru_tipi == "açık_uçlu":
            degerlendirme = llm2_invoke(f"Sorunun doğru cevabı: {sonuc['dogru_cevap']}, kullanıcının cevabı: {sonuc['kullanici_cevap']}, kullanıcının cevabının doğru veya yanlış şeklinde değerlendir. Tek bir kelime ile: 'doğru' veya 'yanlış'")
            dogru_mu = degerlendirme == "doğru"

        # Sonuçları yazdır
        if dogru_mu:
            print("Kullanıcının cevabı: Doğru")
            eksilt_konu_eksiklik_derecesi(dosya_adi, sonuc['ders'], sonuc['konu'])
        else:
            print("Kullanıcının cevabı: Yanlış")
            print(f"Doğru cevap: {sonuc['dogru_cevap']}")
            artir_konu_eksiklik_derecesi(dosya_adi, sonuc['ders'], sonuc['konu'])
        print()  # Boş satır

# Fonksiyonu kullanmak için örnek kullanım
#eksik_konular_sinav_olusturma('eksik_konular.json', 5)  # En yüksek 5 eksiklik derecesine sahip konudan sınav oluşturur

import json
import random

def rastgele_sinav_olusturma(dosya_adi, ders, soru_sayisi):
    # Eksik konular dosyasını oku
    try:
        with open(dosya_adi, 'r', encoding="utf-8") as dosya:
            veriler = json.load(dosya)
    except FileNotFoundError:
        print(f"{dosya_adi} dosyası bulunamadı.")
        return
    except json.JSONDecodeError:
        print("JSON dosyası düzgün okunamadı.")
        return

    # İlgili dersin konularını al
    if ders not in veriler:
        print(f"{ders} dersi eksik konular arasında bulunamadı.")
        return

    konular_ve_dereceler = list(veriler[ders].items())

    # Rastgele konuları seç
    secilen_konular = random.sample(konular_ve_dereceler, min(soru_sayisi, len(konular_ve_dereceler)))

    sorular_ve_cevaplar = []  # Soruları ve cevapları saklamak için liste

    for i, (konu, derece) in enumerate(secilen_konular):
        print(f"\nSoru {i+1} ({ders} - {konu}):")
        # Ders ve konuyu değiştirip yeni soru oluştur
        soru_olusturma(ders, konu, soru_tipi)
        print(soru)  # Soruyu ekrana yazdır
        soru_cevaplama()  # Kullanıcıdan cevap al

        # Soruyu ve cevapları listeye ekle
        sorular_ve_cevaplar.append({
            "ders": ders,
            "konu": konu,
            "soru": soru,
            "dogru_cevap": cevap,
            "kullanici_cevap": kullanici_cevap
        })

    # Sınav bittikten sonra sonuçları kontrol edip yazdır
    print("\n--- Sınav Sonuçları ---\n")
    for i, sonuc in enumerate(sorular_ve_cevaplar):
        print(f"Konu: {sonuc['ders']} - {sonuc['konu']}")
        print(f"Soru {i+1}: {sonuc['soru']}")

        # Doğru cevabı kontrol et
        if soru_tipi == "çoktan_seçmeli":
            dogru_mu = sonuc['dogru_cevap'] == sonuc['kullanici_cevap']
        elif soru_tipi == "açık_uçlu":
            degerlendirme = llm2_invoke(f"Sorunun doğru cevabı: {sonuc['dogru_cevap']}, kullanıcının cevabı: {sonuc['kullanici_cevap']}, kullanıcının cevabının doğru veya yanlış şeklinde değerlendir. Tek bir kelime ile: 'doğru' veya 'yanlış'")
            dogru_mu = degerlendirme == "doğru"

        # Sonuçları yazdır
        if dogru_mu:
            print("Kullanıcının cevabı: Doğru")
            eksilt_konu_eksiklik_derecesi(dosya_adi, sonuc['ders'], sonuc['konu'])
        else:
            print("Kullanıcının cevabı: Yanlış")
            print(f"Doğru cevap: {sonuc['dogru_cevap']}")
            artir_konu_eksiklik_derecesi(dosya_adi, sonuc['ders'], sonuc['konu'])
        print()  # Boş satır

# Fonksiyonu kullanmak için örnek kullanım
#rastgele_sinav_olusturma('eksik_konular.json', 'Coğrafya', 5)  # Biyoloji dersinden 5 rastgele eksik konuyla sınav oluşturur

"""# *Program Olusturucu*"""

import pandas as pd
import json
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

def renk_sec(index, toplam_renkler):
    """ İki aynı hücre için aynı renk, farklı hücre için farklı renk döndürür. """
    renkler = ['ADD8E6', 'E6E6FA', 'FFC0CB']  # Açık mavi, açık mor, açık pembe
    return renkler[index % toplam_renkler]

def autosize_columns(worksheet):
    """Hücre içeriklerine göre sütun genişliğini ayarla."""
    for column_cells in worksheet.columns:
        max_length = 0
        column = column_cells[0].column_letter  # Get the column name
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column].width = adjusted_width

def jsondan_konulari_oku(dosya_adi, ders):
    """JSON dosyasından ders ve konu başlıklarını eksiklik derecelerine göre sıralar."""
    try:
        with open(dosya_adi, 'r', encoding='utf-8') as dosya:
            veriler = json.load(dosya)
            if ders in veriler:
                konular = sorted(veriler[ders].items(), key=lambda x: x[1], reverse=True)
                return [konu for konu, derece in konular]  # Sadece konu isimlerini döndür
            else:
                print(f"{ders} dersi JSON dosyasında bulunamadı.")
                return []
    except FileNotFoundError:
        print(f"{dosya_adi} dosyası bulunamadı.")
        return []
    except json.JSONDecodeError:
        print(f"{dosya_adi} düzgün formatlanmamış.")
        return []

def program_hazirlama_renkli_json(program_suresi, gunluk_saat, json_dosya):
    # Başlangıç tarihi ve gün adları
    gunler = ['Pazartesi', 'Salı', 'Çarşamba', 'Perşembe', 'Cuma', 'Cumartesi', 'Pazar']

    # Saat dilimleri olarak sabah ve akşam
    saat_dilimleri = ['Sabah', 'Akşam']

    # Konuları JSON dosyasından oku ve eksiklik derecelerine göre sırala
    konular = jsondan_konulari_oku(json_dosya, ders)

    if not konular:
        print("Konu bulunamadı.")
        return

    # Konuları sabah ve akşam olmak üzere sırayla programa yerleştirelim
    konu_indeksi = 0
    toplam_konular = len(konular)

    # DataFrame'in kolon başlıkları günler olacak
    data = {gun: ['', ''] for gun in gunler[:program_suresi]}

    # Her gün için sabah ve akşam konu yerleştiriyoruz
    for i in range(program_suresi):
        for saat in range(gunluk_saat):
            # Sabaha ve akşama sırayla konu atıyoruz
            if saat == 0:
                data[gunler[i]][0] = konular[konu_indeksi % toplam_konular]  # Sabah
            elif saat == 1:
                data[gunler[i]][1] = konular[konu_indeksi % toplam_konular]  # Akşam
            konu_indeksi += 1

    # DataFrame oluşturalım
    df = pd.DataFrame(data, index=saat_dilimleri)

    # Excel dosyası ve sayfa oluştur
    dosya_adi = f"calisma_programi_renkli_json.xlsx"
    writer = pd.ExcelWriter(dosya_adi, engine='openpyxl')

    df.to_excel(writer, sheet_name='Program')

    # Excel sayfasına erişelim
    workbook = writer.book
    worksheet = writer.sheets['Program']

    # Hücrelere stil uygulamak için renklendirme seçenekleri
    gun_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Sarı renk (Günler için)
    sabah_aksam_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Yeşil renk (Sabah/Akşam için)

    # Konu hücreleri için renkler
    toplam_renkler = 3  # Açık mavi, açık mor, açık pembe
    renk_indeksi = 0  # Hücrelere sırayla renk atamak için

    # Stil uygulaması - Başlıklar
    for col_num, gun in enumerate(gunler[:program_suresi], start=2):  # Günler için sarı renk
        worksheet.cell(row=1, column=col_num).fill = gun_fill

    # Stil uygulaması - Sabah/Akşam hücreleri
    for row_num, saat in enumerate(saat_dilimleri, start=2):
        worksheet.cell(row=row_num, column=1).fill = sabah_aksam_fill

    # Stil uygulaması - Konu hücreleri
    for row_num, saat in enumerate(saat_dilimleri, start=2):
        for col_num, gun in enumerate(gunler[:program_suresi], start=2):
            icerik = worksheet.cell(row=row_num, column=col_num).value
            if icerik:
                # Eğer aynı konuysa aynı renk olsun
                secilen_renk = renk_sec(renk_indeksi, toplam_renkler)
            else:
                # Farklı konuysa renk değiştirilir
                renk_indeksi += 1
                secilen_renk = renk_sec(renk_indeksi, toplam_renkler)

            konu_fill = PatternFill(start_color=secilen_renk, end_color=secilen_renk, fill_type="solid")
            worksheet.cell(row=row_num, column=col_num).fill = konu_fill

    # Otomatik sütun genişliği ayarı
    autosize_columns(worksheet)

    # Dosyayı kaydet
    writer._save()
    print(f"Çalışma programı {dosya_adi} olarak başarıyla kaydedildi.")

# Örnek kullanım
#program_hazirlama_renkli_json("Biyoloji", program_suresi=7, gunluk_saat=2, json_dosya='eksik_konular.json')

"""# *Retrieval-Augmented-Generation*"""

import cohere

from pinecone import Pinecone
from langchain_openai import ChatOpenAI
from langchain.prompts import ChatPromptTemplate
from langchain_core.output_parsers import StrOutputParser
from sentence_transformers import SentenceTransformer


co = cohere.Client('PWKSJnDago2wzHGA7gbHQldxAd0kIGSWaUcuzGlE')

def get_embedding(sentences):
    if isinstance(sentences, str):  # Convert single sentence to a list
        sentences = [sentences]

    # Call Cohere API with the list of sentences
    response = co.embed(texts=sentences, model="large")

    return response.embeddings


# Initialize Pinecone client
PINECONE_API_KEY = "5a23123c-6de1-48e1-8788-abcc204ebe82"
database = Pinecone(api_key=PINECONE_API_KEY)
INDEX_NAME = "biyoloji"
pinecone_index = database.Index(INDEX_NAME)

# Prompt template
template = """Answer the question based only on the following context:
{context}

Question: {question}
"""

# Chat history
chat_history = []

# Function to interact with the bot
def interact_with_bot(user_query):
    if user_query:
        try:
            # Step 1: Convert user query to a simpler sentence (using llm2_invoke)
            lookup_query = llm2_invoke(f"Kullanıcının aramak istediği bilgiyi bir cümle ile açıkla: {user_query}")

            # Step 2: Retrieve embedding based on lookup query
            xq = get_embedding(lookup_query)
            retrieved = []

            # Step 3: Retrieve relevant information from Pinecone based on embedding
            response = pinecone_index.query(vector=xq, top_k=4, include_metadata=True)

            # Step 4: Extract relevant text from the response
            for match in response["matches"]:
                if 'text' in match["metadata"]:
                    retrieved.append(match["metadata"]["text"])
                else:
                    print(f"'text' key not found in metadata for match: {match}")

            # Step 5: Ensure there is relevant information retrieved
            if retrieved:
                chat_history.append(f"Kullanıcı: {user_query}")

                # Combine context and create a formatted prompt
                combined_context = "\n".join(retrieved)

                # Step 6: Make the prompt explicit about the context and question
                combined_prompt = f"""Cevap verirken sadece aşağıdaki bağlamı kullan:
                Bağlam: {combined_context}

                Soru: {user_query}

                Cevap:
                """

                # Step 7: Invoke the LLM with the combined prompt
                response_text = llm2_invoke(combined_prompt)

                # Add response to chat history
                chat_history.append(f"Bot: {response_text}")

                # Output the result (for testing)
                print(f"Fa'al: {response_text}")

        except Exception as e:
            print(f"Hata: {str(e)}")

def rag():
  user_query = user_input
  interact_with_bot(user_query)

"""# Dosya yuklendiginde konu iceriklerinin parse'*lanmasi*"""

import json
import os

def text_to_json(ders_kazanimlari, dersin_adi, json_data):
    # ". " ile bölerek listeye çevir
    entries = ders_kazanimlari.split("\n")

    # JSON formatına uygun hale getir
    if dersin_adi not in json_data:
        json_data[dersin_adi] = {}  # Yeni bir ders adı ekleniyorsa boş bir konu kümesi oluştur

    for entry in entries:
        if ". " in entry:
            # Her satırda nokta ve boşluk ("1. Hücre yapısı ve işlevleri") formatına uyan bölümü bulalım
            _, content = entry.split(". ", 1)
            json_data[dersin_adi][content.strip()] = 1  # İçeriği anahtar, değeri 1 olarak ekle

    return json_data  # Güncellenmiş JSON verisini döndür

def dosya_yukleme():
    eksik_konular_dosyasi = "eksik_konular.json"

    # Dosya varsa mevcut içeriği yükle, yoksa boş bir dict başlat
    if os.path.exists(eksik_konular_dosyasi):
        with open(eksik_konular_dosyasi, "r", encoding="utf-8") as json_file:
            json_data = json.load(json_file)
    else:
        json_data = {}

    dersin_adi = input("Dersin adı: ")
    sinif = input("Sınıf: ")
    # LLM'den müfredat kazanımlarını alalım
    ders_kazanimlari = llm2_invoke(f"Milli eğitim bakanlığı müfredatına göre {sinif}. sınıf {dersin_adi} dersinin tüm kazanımları nelerdir? Bu kazanımları aralarına boşluk koyarak yaz.")

    # Yeni konuları mevcut JSON'a ekle
    json_data = text_to_json(ders_kazanimlari, dersin_adi, json_data)

    # JSON dosyasına güncellenmiş haliyle kaydet
    with open(eksik_konular_dosyasi, "w", encoding="utf-8") as json_file:
        json.dump(json_data, json_file, ensure_ascii=False, indent=4)
